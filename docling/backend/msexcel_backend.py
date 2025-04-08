import logging
from io import BytesIO
from pathlib import Path
from typing import Any, Union

from docling_core.types.doc import (
    BoundingBox,
    CoordOrigin,
    DoclingDocument,
    DocumentOrigin,
    GroupLabel,
    ImageRef,
    ProvenanceItem,
    Size,
    TableCell,
    TableData,
)
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage
from pydantic import BaseModel
from typing_extensions import override

from docling.backend.abstract_backend import (
    DeclarativeDocumentBackend,
    PaginatedDocumentBackend,
)
from docling.datamodel.base_models import InputFormat
from docling.datamodel.document import InputDocument

_log = logging.getLogger(__name__)


class ExcelCell(BaseModel):
    """Represents an Excel cell.

    Attributes:
        row: The row number of the cell.
        col: The column number of the cell.
        text: The text content of the cell.
        row_span: The number of rows the cell spans.
        col_span: The number of columns the cell spans.
    """

    row: int
    col: int
    text: str
    row_span: int
    col_span: int


class ExcelTable(BaseModel):
    """Represents an Excel table.

    Attributes:
        num_rows: The number of rows in the table.
        num_cols: The number of columns in the table.
        data: The data in the table, represented as a list of ExcelCell objects.
    """

    num_rows: int
    num_cols: int
    data: list[ExcelCell]


class MsExcelDocumentBackend(DeclarativeDocumentBackend, PaginatedDocumentBackend):
    """Backend for parsing Excel workbooks."""

    @override
    def __init__(
        self, in_doc: "InputDocument", path_or_stream: Union[BytesIO, Path]
    ) -> None:
        """Initialize the MsExcelDocumentBackend object.

        Parameters:
            in_doc: The input document object.
            path_or_stream: The path or stream to the Excel file.

        Raises:
            RuntimeError: An error occurred parsing the file.
        """
        super().__init__(in_doc, path_or_stream)

        # Initialise the parents for the hierarchy
        self.max_levels = 10

        self.parents: dict[int, Any] = {}
        for i in range(-1, self.max_levels):
            self.parents[i] = None

        self.workbook = None
        try:
            if isinstance(self.path_or_stream, BytesIO):
                self.workbook = load_workbook(filename=self.path_or_stream)

            elif isinstance(self.path_or_stream, Path):
                self.workbook = load_workbook(filename=str(self.path_or_stream))

            self.valid = self.workbook is not None
        except Exception as e:
            self.valid = False

            raise RuntimeError(
                f"MsExcelDocumentBackend could not load document with hash {self.document_hash}"
            ) from e

    @override
    def is_valid(self) -> bool:
        _log.info(f"valid: {self.valid}")
        return self.valid

    @classmethod
    @override
    def supports_pagination(cls) -> bool:
        return True

    @override
    def page_count(self) -> int:
        if self.is_valid() and self.workbook:
            return len(self.workbook.sheetnames)
        else:
            return 0

    @classmethod
    @override
    def supported_formats(cls) -> set[InputFormat]:
        return {InputFormat.XLSX}

    @override
    def convert(self) -> DoclingDocument:
        """Parse the Excel workbook into a DoclingDocument object.

        Raises:
            RuntimeError: Unable to run the conversion since the backend object failed to
            initialize.

        Returns:
            The DoclingDocument object representing the Excel workbook.
        """
        origin = DocumentOrigin(
            filename=self.file.name or "file.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            binary_hash=self.document_hash,
        )

        doc = DoclingDocument(name=self.file.stem or "file.xlsx", origin=origin)

        if self.is_valid():
            doc = self._convert_workbook(doc)
        else:
            raise RuntimeError(
                f"Cannot convert doc with {self.document_hash} because the backend failed to init."
            )

        return doc

    def _convert_workbook(self, doc: DoclingDocument) -> DoclingDocument:
        """Parse the Excel workbook and attach its structure to a DoclingDocument.

        Args:
            doc: A DoclingDocument object.

        Returns:
            A DoclingDocument object with the parsed items.
        """

        if self.workbook is not None:

            # Iterate over all sheets
            for sheet_name in self.workbook.sheetnames:
                _log.info(f"Processing sheet: {sheet_name}")

                # Access the sheet by name
                sheet = self.workbook[sheet_name]
                idx = self.workbook.index(sheet)
                # TODO: check concept of Size as number of rows and cols
                doc.add_page(page_no=idx + 1, size=Size())

                self.parents[0] = doc.add_group(
                    parent=None,
                    label=GroupLabel.SECTION,
                    name=f"sheet: {sheet_name}",
                )

                doc = self._convert_sheet(doc, sheet)
        else:
            _log.error("Workbook is not initialized.")

        return doc

    def _convert_sheet(self, doc: DoclingDocument, sheet: Worksheet) -> DoclingDocument:
        """Parse an Excel worksheet and attach its structure to a DoclingDocument

        Args:
            doc: The DoclingDocument to be updated.
            sheet: The Excel worksheet to be parsed.

        Returns:
            The updated DoclingDocument.
        """

        doc = self._find_tables_in_sheet(doc, sheet)

        doc = self._find_images_in_sheet(doc, sheet)

        return doc

    def _find_tables_in_sheet(
        self, doc: DoclingDocument, sheet: Worksheet
    ) -> DoclingDocument:
        """Find all tables in an Excel sheet and attach them to a DoclingDocument.

        Args:
            doc: The DoclingDocument to be updated.
            sheet: The Excel worksheet to be parsed.

        Returns:
            The updated DoclingDocument.
        """

        if self.workbook is not None:
            tables = self._find_data_tables(sheet)

            for excel_table in tables:
                num_rows = excel_table.num_rows
                num_cols = excel_table.num_cols

                table_data = TableData(
                    num_rows=num_rows,
                    num_cols=num_cols,
                    table_cells=[],
                )

                for excel_cell in excel_table.data:

                    cell = TableCell(
                        text=excel_cell.text,
                        row_span=excel_cell.row_span,
                        col_span=excel_cell.col_span,
                        start_row_offset_idx=excel_cell.row,
                        end_row_offset_idx=excel_cell.row + excel_cell.row_span,
                        start_col_offset_idx=excel_cell.col,
                        end_col_offset_idx=excel_cell.col + excel_cell.col_span,
                        column_header=excel_cell.row == 0,
                        row_header=False,
                    )
                    table_data.table_cells.append(cell)

                page_no = self.workbook.index(sheet) + 1
                doc.add_table(
                    data=table_data,
                    parent=self.parents[0],
                    prov=ProvenanceItem(
                        page_no=page_no,
                        charspan=(0, 0),
                        bbox=BoundingBox.from_tuple(
                            (0, 0, 0, 0), origin=CoordOrigin.BOTTOMLEFT
                        ),
                    ),
                )

        return doc

    def _find_data_tables(self, sheet: Worksheet) -> list[ExcelTable]:
        """Find all compact rectangular data tables in an Excel worksheet.

        Args:
            sheet: The Excel worksheet to be parsed.

        Returns:
            A list of ExcelTable objects representing the data tables.
        """
        tables: list[ExcelTable] = []  # List to store found tables
        visited: set[tuple[int, int]] = set()  # Track already visited cells

        # Iterate over all cells in the sheet
        for ri, row in enumerate(sheet.iter_rows(values_only=False)):
            for rj, cell in enumerate(row):

                # Skip empty or already visited cells
                if cell.value is None or (ri, rj) in visited:
                    continue

                # If the cell starts a new table, find its bounds
                table_bounds, visited_cells = self._find_table_bounds(sheet, ri, rj)

                visited.update(visited_cells)  # Mark these cells as visited
                tables.append(table_bounds)

        return tables

    def _find_table_bounds(
        self,
        sheet: Worksheet,
        start_row: int,
        start_col: int,
    ) -> tuple[ExcelTable, set[tuple[int, int]]]:
        """Determine the bounds of a compact rectangular table.

        Args:
            sheet: The Excel worksheet to be parsed.
            start_row: The row number of the starting cell.
            start_col: The column number of the starting cell.

        Returns:
            A tuple with an Excel table and a set of cell coordinates.
        """
        _log.debug("find_table_bounds")

        max_row = self._find_table_bottom(sheet, start_row, start_col)
        max_col = self._find_table_right(sheet, start_row, start_col)

        # Collect the data within the bounds
        data = []
        visited_cells: set[tuple[int, int]] = set()
        for ri in range(start_row, max_row + 1):
            for rj in range(start_col, max_col + 1):

                cell = sheet.cell(row=ri + 1, column=rj + 1)  # 1-based indexing

                # Check if the cell belongs to a merged range
                row_span = 1
                col_span = 1

                # _log.info(sheet.merged_cells.ranges)
                for merged_range in sheet.merged_cells.ranges:

                    if (
                        merged_range.min_row <= ri + 1
                        and ri + 1 <= merged_range.max_row
                        and merged_range.min_col <= rj + 1
                        and rj + 1 <= merged_range.max_col
                    ):

                        row_span = merged_range.max_row - merged_range.min_row + 1
                        col_span = merged_range.max_col - merged_range.min_col + 1
                        break

                if (ri, rj) not in visited_cells:
                    data.append(
                        ExcelCell(
                            row=ri - start_row,
                            col=rj - start_col,
                            text=str(cell.value),
                            row_span=row_span,
                            col_span=col_span,
                        )
                    )
                    # _log.info(f"cell: {ri}, {rj} -> {ri - start_row}, {rj - start_col}, {row_span}, {col_span}: {str(cell.value)}")

                    # Mark all cells in the span as visited
                    for span_row in range(ri, ri + row_span):
                        for span_col in range(rj, rj + col_span):
                            visited_cells.add((span_row, span_col))

        return (
            ExcelTable(
                num_rows=max_row + 1 - start_row,
                num_cols=max_col + 1 - start_col,
                data=data,
            ),
            visited_cells,
        )

    def _find_table_bottom(
        self, sheet: Worksheet, start_row: int, start_col: int
    ) -> int:
        """Find the bottom boundary of a table.

        Args:
            sheet: The Excel worksheet to be parsed.
            start_row: The starting row of the table.
            start_col: The starting column of the table.

        Returns:
            The row index representing the bottom boundary of the table.
        """
        max_row: int = start_row

        while max_row < sheet.max_row - 1:
            # Get the cell value or check if it is part of a merged cell
            cell = sheet.cell(row=max_row + 2, column=start_col + 1)

            # Check if the cell is part of a merged range
            merged_range = next(
                (mr for mr in sheet.merged_cells.ranges if cell.coordinate in mr),
                None,
            )

            if cell.value is None and not merged_range:
                break  # Stop if the cell is empty and not merged

            # Expand max_row to include the merged range if applicable
            if merged_range:
                max_row = max(max_row, merged_range.max_row - 1)
            else:
                max_row += 1

        return max_row

    def _find_table_right(
        self, sheet: Worksheet, start_row: int, start_col: int
    ) -> int:
        """Find the right boundary of a table.

        Args:
            sheet: The Excel worksheet to be parsed.
            start_row: The starting row of the table.
            start_col: The starting column of the table.

        Returns:
            The column index representing the right boundary of the table."
        """
        max_col: int = start_col

        while max_col < sheet.max_column - 1:
            # Get the cell value or check if it is part of a merged cell
            cell = sheet.cell(row=start_row + 1, column=max_col + 2)

            # Check if the cell is part of a merged range
            merged_range = next(
                (mr for mr in sheet.merged_cells.ranges if cell.coordinate in mr),
                None,
            )

            if cell.value is None and not merged_range:
                break  # Stop if the cell is empty and not merged

            # Expand max_col to include the merged range if applicable
            if merged_range:
                max_col = max(max_col, merged_range.max_col - 1)
            else:
                max_col += 1

        return max_col

    def _find_images_in_sheet(
        self, doc: DoclingDocument, sheet: Worksheet
    ) -> DoclingDocument:
        """Find images in the Excel sheet and attach them to the DoclingDocument.

        Args:
            doc: The DoclingDocument to be updated.
            sheet: The Excel worksheet to be parsed.

        Returns:
            The updated DoclingDocument.
        """

        if self.workbook is not None:
            # Iterate over byte images in the sheet
            for image in sheet._images:  # type: ignore[attr-defined]
                try:
                    pil_image = PILImage.open(image.ref)
                    page_no = self.workbook.index(sheet) + 1
                    doc.add_picture(
                        parent=self.parents[0],
                        image=ImageRef.from_pil(image=pil_image, dpi=72),
                        caption=None,
                        prov=ProvenanceItem(
                            page_no=page_no,
                            charspan=(0, 0),
                            bbox=BoundingBox.from_tuple(
                                (0, 0, 0, 0), origin=CoordOrigin.TOPLEFT
                            ),
                        ),
                    )
                except:
                    _log.error("could not extract the image from excel sheets")

        return doc
