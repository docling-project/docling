# 엑셀(xlsx/xlsm)·csv 직접 처리 헬퍼 (이슈 #288 / KOBCI-183)
#
# intelligent_processor 가 PDF 변환 없이 xlsx 를 직접 처리하기 위한 헬퍼 모듈.
# facade 가 아니라 converters 하위 모듈이며, intelligent_processor → 이 모듈의 단방향 의존만 갖는다.
#
# 두 가지 처리 방식을 제공한다.
#   - build_docling_document(): docling MsExcelDocumentBackend 로 DoclingDocument 생성(시트=1페이지).
#       → 기존 청킹/벡터 파이프라인을 그대로 태운다. PDF 변환 시 한 행이 페이지 경계로 쪼개지는
#         논리 오류를 원천 제거한다.
#   - build_tabular_vectors(): 데이터 행마다 1청크, 컬럼 헤더→셀 값을 메타데이터로 부여한다.
#       openpyxl 로 병합셀을 unmerge + forward-fill 하여 병합 헤더 유실을 방지한다.
#
# docling 관련 import 는 build_docling_document() 내부에서만 수행한다(로컬 vendored docling import
# 충돌 회피 + tabular 전용 사용 시 docling 미로딩).
from __future__ import annotations

import hashlib
import json
import logging
import os
import re
from datetime import datetime
from typing import Any, Optional

from pydantic import BaseModel

_log = logging.getLogger(__name__)

XLSX_EXTS = {".xlsx", ".xlsm"}
CSV_EXTS = {".csv"}

# Weaviate property name 제약(GraphQL 표준): /[_A-Za-z][_0-9A-Za-z]*/
# 한글 등 비-ASCII 헤더를 메타데이터 KEY 로 쓰면 grpc 에러가 나므로 키 후보에서 제외한다.
_VALID_KEY_RE = re.compile(r"^[_A-Za-z][_0-9A-Za-z]*$")

# 예약 필드(컬럼 헤더가 이 이름과 충돌하면 메타 키로 쓰지 않는다)
_RESERVED_FIELDS = {
    "text", "n_char", "n_word", "n_line", "i_page", "e_page",
    "i_chunk_on_page", "n_chunk_of_page", "i_chunk_on_doc", "n_chunk_of_doc",
    "n_page", "reg_date", "chunk_bboxes", "media_files", "column_map",
}


class GenOSVectorMeta(BaseModel):
    class Config:
        extra = "allow"

    text: Optional[str] = None
    n_char: Optional[int] = None
    n_word: Optional[int] = None
    n_line: Optional[int] = None
    i_page: Optional[int] = None
    e_page: Optional[int] = None
    i_chunk_on_page: Optional[int] = None
    n_chunk_of_page: Optional[int] = None
    i_chunk_on_doc: Optional[int] = None
    n_chunk_of_doc: Optional[int] = None
    n_page: Optional[int] = None
    reg_date: Optional[str] = None
    chunk_bboxes: Optional[str] = None
    media_files: Optional[str] = None


# ------------------------------------------------------------------ #
# 공통 헬퍼                                                            #
# ------------------------------------------------------------------ #
def is_xlsx_like(file_path: str) -> bool:
    return os.path.splitext(file_path)[1].lower() in XLSX_EXTS


def is_csv(file_path: str) -> bool:
    return os.path.splitext(file_path)[1].lower() in CSV_EXTS


def _safe_utf8(v: str) -> str:
    return v.encode("utf-8", errors="ignore").decode("utf-8")


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return _safe_utf8(v) if isinstance(v, str) else str(v)


def _row_to_pipe(values: list) -> str:
    cells = " | ".join(_cell_str(v) for v in values)
    return f"| {cells} |"


def _is_valid_key(key: str) -> bool:
    return bool(key) and key not in _RESERVED_FIELDS and bool(_VALID_KEY_RE.match(key))


def _stable_key(name: str) -> str:
    """헤더 텍스트 → Weaviate property 필터에 쓸 안정적 ASCII 키.

    같은 헤더 텍스트는 어느 파일에서든 같은 키가 되도록 결정적으로 생성한다.
      - ASCII 규칙(_is_valid_key) 통과(영문 헤더 등) → 그대로(가독성 유지)
      - 그 외(한글·공백·기호 등) → 'field_' + sha256(header)[:8]
      - 빈 헤더 → '' (호출부에서 위치기반 col_N 로 폴백)
    """
    name = (name or "").strip()
    if not name:
        return ""
    if _is_valid_key(name):
        return name
    return "field_" + hashlib.sha256(name.encode("utf-8")).hexdigest()[:8]


# ------------------------------------------------------------------ #
# docling 모드                                                         #
# ------------------------------------------------------------------ #
_docling_converter = None  # lazy singleton


def _get_docling_converter():
    """xlsx/csv 전용 DocumentConverter 를 생성·캐시한다(PDF 컨버터와 분리)."""
    global _docling_converter
    if _docling_converter is None:
        from docling.datamodel.base_models import InputFormat
        from docling.document_converter import (
            CsvFormatOption,
            DocumentConverter,
            ExcelFormatOption,
        )

        _docling_converter = DocumentConverter(
            allowed_formats=[InputFormat.XLSX, InputFormat.CSV],
            format_options={
                InputFormat.XLSX: ExcelFormatOption(),
                InputFormat.CSV: CsvFormatOption(),
            },
        )
    return _docling_converter


def build_docling_document(file_path: str, *, save_images: bool = False):
    """xlsx/csv 를 docling MsExcel/Csv 백엔드로 변환해 DoclingDocument 를 반환한다.

    SimplePipeline 은 save_images 옵션을 사용하지 않는다(엑셀에 렌더 이미지 없음).
    변환 실패 시 예외를 그대로 올린다(호출부에서 GenosServiceException 으로 매핑).
    """
    converter = _get_docling_converter()
    conv_result = converter.convert(file_path, raises_on_error=True)
    return conv_result.document


# ------------------------------------------------------------------ #
# tabular 모드                                                         #
# ------------------------------------------------------------------ #
def _detect_encoding(file_path: str, fallback: Optional[str]) -> str:
    if fallback:
        return fallback
    try:
        import chardet

        with open(file_path, "rb") as f:
            raw = f.read(10_000)
        return chardet.detect(raw).get("encoding") or "utf-8"
    except Exception:
        return "utf-8"


def _csv_rows(file_path: str, encoding: Optional[str]) -> list[list[str]]:
    import csv as _csv

    enc = _detect_encoding(file_path, encoding)
    rows: list[list[str]] = []
    with open(file_path, "r", encoding=enc, errors="ignore", newline="") as f:
        for row in _csv.reader(f):
            rows.append([_cell_str(c) for c in row])
    return rows


def _sheet_rows_unmerged(ws) -> list[list[str]]:
    """병합셀을 unmerge + forward-fill 한 뒤 2D 문자열 행 목록을 반환한다."""
    for rng in list(ws.merged_cells.ranges):
        top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
        ws.unmerge_cells(str(rng))
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                ws.cell(row=r, column=c).value = top_left

    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_str(v) for v in row])
    return rows


def _load_sheets(file_path: str, encoding: Optional[str]) -> dict[str, list[list[str]]]:
    if is_csv(file_path):
        return {"table_1": _csv_rows(file_path, encoding)}

    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    return {name: _sheet_rows_unmerged(wb[name]) for name in wb.sheetnames}


def load_sheets(file_path: str, *, encoding: Optional[str] = None) -> dict[str, list[list[str]]]:
    """xlsx/csv 를 시트명 → 2D 문자열 행 목록으로 로드한다(공개 API).

    xlsx 는 병합셀을 unmerge + forward-fill 하여 병합 헤더/그룹 값 유실을 막는다.
    csv 는 인코딩 자동감지(또는 encoding) 후 단일 시트("table_1")로 반환한다.
    벡터/HTML 등 출력 형태에 독립적인 중립 표현이라 여러 facade 에서 재사용한다.
    """
    return _load_sheets(file_path, encoding)


def _is_empty_row(row: list[str]) -> bool:
    return all(not c.strip() for c in row)


def _sheet_grid_and_merges(ws) -> tuple[list[list[str]], list[tuple[int, int, int, int]]]:
    """시트를 (2D 문자열 행, 병합범위 목록)으로 로드한다.

    병합범위는 unmerge 이전에 0-based `(r0, c0, r1, c1)`(포함) 로 캡처한 뒤, 값은 unmerge +
    forward-fill 하여 반환한다. 병합정보는 멀티헤더(제목/계층) 판정에 사용한다.
    """
    merges = [
        (rng.min_row - 1, rng.min_col - 1, rng.max_row - 1, rng.max_col - 1)
        for rng in ws.merged_cells.ranges
    ]
    for rng in list(ws.merged_cells.ranges):
        top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
        ws.unmerge_cells(str(rng))
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                ws.cell(row=r, column=c).value = top_left
    rows = [[_cell_str(v) for v in row] for row in ws.iter_rows(values_only=True)]
    return rows, merges


def _load_sheets_with_merges(
    file_path: str, encoding: Optional[str] = None
) -> dict[str, tuple[list[list[str]], list[tuple[int, int, int, int]]]]:
    if is_csv(file_path):
        return {"table_1": (_csv_rows(file_path, encoding), [])}

    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    return {name: _sheet_grid_and_merges(wb[name]) for name in wb.sheetnames}


def _split_blocks(
    rows: list[list[str]],
    merges: list[tuple[int, int, int, int]],
    multi_table: bool,
) -> list[tuple[list[list[str]], list[tuple[int, int, int, int]]]]:
    """시트를 표 블록으로 분리한다. multi_table 이면 완전 빈 행 run 을 구분자로, 아니면 1블록."""
    if not multi_table:
        return [(rows, merges)]

    spans: list[tuple[int, int]] = []
    start: Optional[int] = None
    for i, r in enumerate(rows):
        if not _is_empty_row(r):
            if start is None:
                start = i
        elif start is not None:
            spans.append((start, i))
            start = None
    if start is not None:
        spans.append((start, len(rows)))

    result = []
    for a, b in spans:
        brows = rows[a:b]
        bmerges = [
            (r0 - a, c0, r1 - a, c1)
            for (r0, c0, r1, c1) in merges
            if r0 >= a and r1 < b
        ]
        result.append((brows, bmerges))
    return result


def _detect_header(
    brows: list[list[str]],
    bmerges: list[tuple[int, int, int, int]],
    used_cols: list[int],
    header_row_override: int,
) -> tuple[list[int], list[int], int, int]:
    """블록의 헤더 구조를 자동 판정한다(병합 기반).

    반환: (title_rows, group_rows, leaf_idx, data_start)
      - title_rows: 단일 수평 병합만으로 구성된 제목행(키 제외, 컨텍스트).
      - group_rows: 수평 병합이 있는 계층 헤더행(상위 레벨).
      - leaf_idx: 컬럼명행(수평 병합 없는 첫 행).
    header_row_override>0 이면 자동판정 대신 그 인덱스를 leaf 로 강제(레거시 단일헤더).
    """
    n = len(brows)
    if header_row_override and header_row_override > 0:
        leaf = min(header_row_override, n - 1)
        return [], [], leaf, leaf + 1

    hmerge_rows: dict[int, list[tuple[int, int]]] = {}
    for (r0, c0, r1, c1) in bmerges:
        if c1 > c0:  # 수평 span
            for r in range(r0, r1 + 1):
                hmerge_rows.setdefault(r, []).append((c0, c1))

    title_rows: list[int] = []
    group_rows: list[int] = []
    leaf_idx: Optional[int] = None
    i = 0
    while i < n:
        row = brows[i]
        ne = [c for c in used_cols if c < len(row) and row[c].strip()]
        if not ne:
            i += 1
            continue
        hms = hmerge_rows.get(i, [])
        if len(hms) == 1:
            mc0, mc1 = hms[0]
            if all(mc0 <= c <= mc1 for c in ne):  # 단일 병합이 모든 내용 커버 → 제목행
                title_rows.append(i)
                i += 1
                continue
        if hms:  # 수평 병합 있음 → 계층 헤더행
            group_rows.append(i)
            i += 1
            continue
        leaf_idx = i  # 병합 없는 첫 행 → 컬럼명행
        break

    if leaf_idx is None:
        if i < n:
            leaf_idx = i
        elif group_rows:
            leaf_idx = group_rows.pop()
        elif title_rows:
            leaf_idx = title_rows.pop()
        else:
            leaf_idx = 0
    return title_rows, group_rows, leaf_idx, leaf_idx + 1


def _row_first_text(row: list[str]) -> str:
    for v in row:
        if v.strip():
            return v.strip()
    return ""


def load_tables(
    file_path: str,
    *,
    header_row: int = 0,
    multi_table: bool = False,
) -> list[dict]:
    """xlsx/csv 를 표 블록 목록으로 감지·추출한다(출력 형태에 중립 — 벡터/HTML 공용).

    각 블록 dict:
      - sheet_name, sheet_index(1-based)
      - title: 제목행 텍스트(컨텍스트; 키 아님)
      - headers: used_cols 순서의 flatten 헤더명(계층 병합은 `상위_하위`)
      - data_rows: [[used_cols 순서의 셀 값 ...], ...] (비어있지 않은 데이터 행)

    멀티헤더 자동 판정(_detect_header)·복수표 분리(_split_blocks) 적용.
    header_row>0 이면 레거시 단일헤더(그 인덱스) 강제.
    """
    sheets = _load_sheets_with_merges(file_path)
    tables: list[dict] = []

    for sheet_idx, (sheet_name, (rows, merges)) in enumerate(sheets.items(), start=1):
        for brows, bmerges in _split_blocks(rows, merges, multi_table):
            if not brows:
                continue
            used_cols = sorted({c for r in brows for c in range(len(r)) if r[c].strip()})
            if not used_cols:
                continue

            title_rows, group_rows, leaf_idx, data_start = _detect_header(
                brows, bmerges, used_cols, header_row
            )
            data_rows = [r for r in brows[data_start:] if not _is_empty_row(r)]
            if not data_rows:
                continue

            # 컬럼별 헤더명(계층 flatten). title 은 제외.
            headers: list[str] = []
            for c in used_cols:
                parts = []
                for g in group_rows:
                    v = brows[g][c].strip() if c < len(brows[g]) else ""
                    if v:
                        parts.append(v)
                lv = brows[leaf_idx][c].strip() if leaf_idx < len(brows) and c < len(brows[leaf_idx]) else ""
                if lv:
                    parts.append(lv)
                headers.append("_".join(parts))

            title_text = " / ".join(
                dict.fromkeys(
                    _row_first_text(brows[t]) for t in title_rows if _row_first_text(brows[t])
                )
            )
            values_rows = [[(r[c] if c < len(r) else "") for c in used_cols] for r in data_rows]
            tables.append({
                "sheet_name": sheet_name,
                "sheet_index": sheet_idx,
                "title": title_text,
                "headers": headers,
                "data_rows": values_rows,
            })

    return tables


def build_tabular_vectors(
    file_path: str,
    *,
    header_row: int = 0,
    multi_table: bool = False,
    reg_date: Optional[str] = None,
) -> list[GenOSVectorMeta]:
    """데이터 행마다 1벡터(GenOSVectorMeta). 표 감지는 load_tables 에 위임하고, 여기서는 벡터 메타
    계층(헤더 기반 안정 키·column_map·page/chunk 필드)만 담당한다.

    각 컬럼 값을 최상단 스칼라 property 로 부여(Weaviate where 필터 가능). Weaviate 키 규칙에 안 맞는
    헤더(한글 등)는 `_stable_key` 로 `field_<hash>` alias, 원본명은 `column_map`(JSON) 에 보존한다.
    """
    reg_date = reg_date or (datetime.now().isoformat(timespec="seconds") + "Z")
    tables = load_tables(file_path, header_row=header_row, multi_table=multi_table)

    n_chunk_of_doc = sum(len(t["data_rows"]) for t in tables)
    if n_chunk_of_doc == 0:
        _log.warning(f"[xlsx] tabular 청크 없음: {file_path}")
        return []
    n_page = max((t["sheet_index"] for t in tables), default=0)

    vectors: list[GenOSVectorMeta] = []
    chunk_doc_idx = 0

    for table in tables:
        sheet_name = table["sheet_name"]
        sheet_idx = table["sheet_index"]
        title_text = table["title"]
        headers = table["headers"]
        data_rows = table["data_rows"]

        # 헤더 기반 안정 키(같은 헤더=같은 키) + column_map(원본명 보존). 표 내 충돌만 suffix.
        keys: list[str] = []
        column_map: dict[str, str] = {}
        used_keys: set[str] = set()
        for i, name in enumerate(headers):
            key = _stable_key(name) or f"col_{i + 1}"
            base, k = key, 2
            while key in used_keys:
                key = f"{base}_{k}"
                k += 1
            used_keys.add(key)
            keys.append(key)
            if name:
                column_map[key] = name
        column_map_json = json.dumps(column_map, ensure_ascii=False)
        header_line = _row_to_pipe(headers)

        for row_idx, values in enumerate(data_rows):
            text_parts = [f"시트명: {sheet_name}"]
            if title_text:
                text_parts.append(title_text)
            text_parts.append(header_line)
            text_parts.append(_row_to_pipe(values))
            text = "\n".join(text_parts)

            row_fields = {keys[i]: values[i] for i in range(len(keys))}
            vectors.append(
                GenOSVectorMeta.model_validate(
                    {
                        **row_fields,
                        "text": text,
                        "n_char": len(text),
                        "n_word": len(text.split()),
                        "n_line": len(text.splitlines()),
                        "i_page": sheet_idx,
                        "e_page": sheet_idx,
                        "i_chunk_on_page": row_idx,
                        "n_chunk_of_page": len(data_rows),
                        "i_chunk_on_doc": chunk_doc_idx,
                        "n_chunk_of_doc": n_chunk_of_doc,
                        "n_page": n_page,
                        "reg_date": reg_date,
                        "chunk_bboxes": ".",
                        "media_files": ".",
                        "column_map": column_map_json,
                    }
                )
            )
            chunk_doc_idx += 1

    return vectors
